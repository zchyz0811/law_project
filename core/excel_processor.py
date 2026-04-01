from openpyxl import load_workbook
from core.template_engine import TemplateEngine, VariableOccurrence
from typing import Dict, List, Tuple


class ExcelProcessor:
    def __init__(self):
        self.engine = TemplateEngine()

    def extract_text(self, path: str) -> str:
        """提取Excel的所有文本"""
        wb = load_workbook(path)
        all_text = []

        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        all_text.append(str(cell.value))

        return "\n".join(all_text)

    def extract_variables(self, path: str):
        """从已有模板中提取变量"""
        text = self.extract_text(path)
        return self.engine.extract_variables(text)

    def apply_replacements_by_positions(self, source_path: str, output_path: str,
                                        replacement_list: List[Tuple[int, int, str]]):
        """按位置执行替换列表"""
        wb = load_workbook(source_path)

        # 构建单元格偏移映射
        segments = []  # [(cell, offset, length)]
        offset = 0

        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    text = str(cell.value) if cell.value else ""
                    segments.append((cell, offset, len(text)))
                    offset += len(text) + 1

        # 按单元格分组
        cell_replacements: Dict[int, List[Tuple[int, int, str]]] = {}
        for start, end, new_text in replacement_list:
            for idx, (cell, seg_offset, seg_len) in enumerate(segments):
                seg_end = seg_offset + seg_len
                if start >= seg_offset and end <= seg_end:
                    local_start = start - seg_offset
                    local_end = end - seg_offset
                    cell_replacements.setdefault(idx, []).append(
                        (local_start, local_end, new_text)
                    )
                    break

        # 对每个单元格执行替换
        for idx, reps in cell_replacements.items():
            cell = segments[idx][0]
            if not cell.value or not isinstance(cell.value, str):
                continue
            text = cell.value
            for local_start, local_end, new_text in sorted(reps, key=lambda r: r[0], reverse=True):
                text = text[:local_start] + new_text + text[local_end:]
            cell.value = text

        wb.save(output_path)

    def create_template(self, original_path: str, output_path: str,
                        replacements: Dict[str, str]):
        """将原始Excel转换为模板（全局替换，保留向后兼容）"""
        wb = load_workbook(original_path)

        sorted_items = sorted(replacements.items(), key=lambda x: len(x[0]),
                              reverse=True)

        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        new_value = cell.value
                        for original, placeholder in sorted_items:
                            new_value = new_value.replace(original, placeholder)
                        if new_value != cell.value:
                            cell.value = new_value

        wb.save(output_path)

    def fill_and_save(self, template_path: str, output_path: str,
                      values: Dict[str, str]):
        """填充模板并保存"""
        wb = load_workbook(template_path)

        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell.value = self.engine.fill(cell.value, values)

        wb.save(output_path)
